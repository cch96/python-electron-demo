# coding:utf-8
import win32api, win32con, win32gui
from win32api import GetSystemMetrics
import openpyxl
from openpyxl import load_workbook
import pyperclip
import time
import traceback
import random

import qzm_tools


VK_CODE = {
    'a':0x41,
    'ctrl':0x11,
    'c':0x43,
    'v':0x56,
    'enter':0x0D,
    'F5':0x74
}

alert_height=0
elements = {}
bot_height = 0

def get_element(element_name):
    #  返回缓存的元素位置，若没有则截图对比获取位置
    global elements
    if element_name not in elements:
        qzm_tools.ImgFinder.window_capture()
        time.sleep(2)
        btnfinder =  qzm_tools.ImgFinder(element_name)
        elements[element_name] = btnfinder.get_element()
    return elements[element_name]


def keybord_select_all():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('a')
    keybord_input('c')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)
    
def keybord_input(key):
    win32api.keybd_event(VK_CODE[key], 0, 0, 0)
    win32api.keybd_event(VK_CODE[key], 0, win32con.KEYEVENTF_KEYUP, 0)

def keybord_paste():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('v')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)
    keybord_input('enter')

def get_top_windows():
    windList = []
    win32gui.EnumWindows(lambda hWnd, param: param.append(hWnd), windList)
    return windList

def set_wind_pos():
    global hwnd
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('- 工作台') > -1:
            hwnd = wind
    # 置顶并全屏对话框
    win32gui.BringWindowToTop(hwnd)
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(hwnd)

def clear_wind_post():
    win32gui.SetWindowPos(hwnd, win32con.HWND_BOTTOM
                              , GetSystemMetrics(0)-1280
                              , 0, 1280, 680
                              , win32con.SWP_NOACTIVATE| win32con.SWP_NOOWNERZORDER|win32con.SWP_SHOWWINDOW)


def add_new_ques(ques):
    #点击输入框
    print('点击输入框')
    win32api.SetCursorPos(
        [int(GetSystemMetrics(0)/2), int(GetSystemMetrics(1)/2-49)])
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.3)
    #输入内容
    print('输入内容')
    pyperclip.copy(ques)
    keybord_paste()
    time.sleep(0.2)
    #点击保存
    print('点击保存')
    win32api.SetCursorPos(get_element('save_custom_qus.png'))
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.8)
    #检测是否错误
    keybord_select_all()
    time.sleep(0.2)
    res = pyperclip.paste()
    time.sleep(0.2)
    win32api.mouse_event(win32con.MOUSEEVENTF_MIDDLEDOWN,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_MIDDLEUP,0,0,0)
    if res.find('错误提醒') > -1:
        print('重复，不管')
        # 为避免提示重复问题，页面点两下
        win32api.SetCursorPos([int(GetSystemMetrics(
            0) - random.randint(100, 200)), int(GetSystemMetrics(1) - 150)])
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(0.5)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(0.5)
    elif res.find("操作确认") > -1:
        print('点击确定')
        #点击确定
        try:
            # 可能没有强行确认按钮
            win32api.SetCursorPos(get_element('force_confirm.png'))
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
            time.sleep(0.5)
        except Exception:
            print('没有强行确定按键')
        print('点击确定')
    else:
        # 取消全选
        win32api.SetCursorPos([int(GetSystemMetrics(0) - 564), 237])
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)

def create_new_item(ques):
    #点击添加按钮
    global alert_height
    print('点击添加按钮')
    win32api.SetCursorPos(get_element('add_custom.png'))
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.5)

    add_new_ques(ques)
    #下拉
    print('下拉: 40')
    win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL,0,0,-55)
    time.sleep(0.5)
    print('已添加新词条\n')

def add_ques(ques):
    #点击添加问法
    global alert_height
    print('点击添加问法')
    qzm_tools.ImgFinder.window_capture()
    time.sleep(0.8)
    btn = qzm_tools.ImgFinder('add_qus.png')
    win32api.SetCursorPos(btn.get_element())
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.8)
    #点击输入框
    add_new_ques(ques)
    # print('点击输入框')
    # win32api.SetCursorPos([int(GetSystemMetrics(0)/2), int(GetSystemMetrics(1)/2-49)])
    # time.sleep(0.2)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    # time.sleep(0.3)
    #
    # #输入内容
    # print('输入内容')
    # pyperclip.copy(ques)
    # keybord_paste()
    # time.sleep(0.5)
    #
    # #点击保存
    # print('点击保存')
    # win32api.SetCursorPos(
    #     [int(GetSystemMetrics(0)/2 + 400), int(GetSystemMetrics(1)/2+350)])
    # time.sleep(0.2)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    # time.sleep(0.3)
    #下拉
    print('下拉: 53')
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL,0,0,-51)
    time.sleep(0.5)
    print('已添加问法\n')

def get_excel_content():
    result_list = []
    excel = load_workbook('QA表.xlsx')
    wb = excel.active
    start = None
    end = None
    for i in range(3, wb.max_row+1):
        if wb.cell(row=i, column=4).value != None:
            temp = get_ques_list(wb.cell(row=i, column=4).value)
            temp = temp.split('\n')
            temp2 = []
            for item in temp:
                if item != '':
                    temp2.append(item)
            result_list.append(temp2)

        #if len(result_list) == 2:
        #    break
    return result_list


def get_ques_list(ques):
    ques.strip()
    ques.strip('/')
    ques = ques.replace('/', '\n')
    return ques

def turn2morequs():
    global bot_height
    keybord_input('F5')
    time.sleep(4)
    # 获得任务栏高度
    for wind in get_top_windows():
        if win32gui.GetClassName(wind) == 'Shell_TrayWnd':
            bot_height = win32gui.GetWindowRect(
                wind)[3] - win32gui.GetWindowRect(wind)[1]
    # 关闭弹出页面
    w_height = GetSystemMetrics(1)-100-bot_height
    height = int(w_height/2)-140+100
    width = int((GetSystemMetrics(0))/2)+285
    win32api.SetCursorPos([width, height])
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    # 截取屏幕分析按键位置
    qzm_tools.ImgFinder.window_capture()
    # 点击更多问题，方便将来导入
    win32api.SetCursorPos(get_element('more_qus.png'))
    # win32api.SetCursorPos([880, 230+alert_height])

    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.8)

def main():
    set_wind_pos()
    result_list = get_excel_content()
    turn2morequs()
    num = 0
    for a in range(0, len(result_list)):
        num += 1
        added_num = 0
        for i in range(0, len(result_list[a])):
            added_num += 1
            if i == 0:
                create_new_item(result_list[a][i])
            else:
                add_ques(result_list[a][i])

        #每完成一条后，回到顶部
        print('下拉: 10000')
        time.sleep(0.5)
        win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL,0,0,10000)
        time.sleep(1)
        if num == 10:
            turn2morequs()
            num = 0

        # if a == 2:
        #     break
    clear_wind_post()

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        traceback.print_exc()
    input()





