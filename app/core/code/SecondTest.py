# coding:utf-8
import win32api
import win32con
import win32gui
from win32api import GetSystemMetrics
import openpyxl
from openpyxl import load_workbook
import pyperclip
import time
import random
import os
import pythoncom
import win32com.client
import traceback

import qzm_tools

VK_CODE = {
    'a': 0x41,
    'ctrl': 0x11,
    'c': 0x43,
    'v': 0x56,
    'enter': 0x0D,
    'F5': 0x74,
    'F12': 0x7B
}

timestamp = str(time.time())
chat_wnd = None
hwndList = None
parent_hwnd = None
bot_height = 0
ans_hwnd = None
confirm_hwnd = None
alert_height = 20

elements = {}

def get_element(element_name):
    #  返回缓存的元素位置，若没有则截图对比获取位置
    global elements
    if element_name not in elements:
        qzm_tools.ImgFinder.window_capture()
        time.sleep(2)
        btnfinder =  qzm_tools.ImgFinder(element_name)
        elements[element_name] = btnfinder.get_element()
    return elements[element_name]


def find_ans_input():
    global ans_hwnd
    global confirm_hwnd
    par_hwnd = None
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('答案编辑框') > -1:
            par_hwnd = wind
            break
    children_winds = []
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), children_winds)
    for cwind in children_winds:
        if win32gui.GetClassName(cwind).find('RichEditComponent') > -1:
            ans_hwnd = cwind
        if win32gui.GetWindowText(cwind).find('确') > -1:
            confirm_hwnd = cwind
    return ans_hwnd


def check_chat_window(hWnd):
    if not hWnd:
        return Falsse
    if '- 工作台' in win32gui.GetWindowText(hWnd):
        return True


def check_bot_window(hWnd):
    if not hWnd:
        return Falsse
    if 'MSTaskListWClass' in win32gui.GetClassName(hWnd):
        return True


def get_top_windows():
    windList = []
    win32gui.EnumWindows(lambda hWnd, param: param.append(hWnd), windList)
    return windList


def find_target_hwnd(parent_hwnd):
    child_list = []
    win32gui.EnumWindows(lambda hwnd, param: param.append(hwnd), child_list)
    for wind in child_list:
        if check_chat_window(wind):
            first_wind = wind
            break
    return first_wind


def get_edit_hwnd(hwnd):
    c1 = None
    c1_list = []
    temp = []
    win32gui.EnumChildWindows(
        hwnd, lambda hWnd, param: param.append(hWnd), c1_list)
    for item in c1_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c2 = temp[1]
    temp = []
    c2_list = []

    win32gui.EnumChildWindows(
        c2, lambda hWnd, param: param.append(hWnd), c2_list)
    for item in c2_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c3 = temp[1]
    temp = []
    c3_list = []

    win32gui.EnumChildWindows(
        c3, lambda hWnd, param: param.append(hWnd), c3_list)
    for item in c3_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c4 = temp[0]
    temp = []
    c4_list = []

    win32gui.EnumChildWindows(
        c4, lambda hWnd, param: param.append(hWnd), c4_list)
    for item in c4_list:
        if win32gui.GetClassName(item) == 'StandardWindow':
            temp.append(item)
    c5 = temp[0]
    temp = []
    c5_list = []
    win32gui.EnumChildWindows(
        c5, lambda hWnd, param: param.append(hWnd), c5_list)
    for item in c5_list:
        if win32gui.GetClassName(item) == 'EditComponent':
            temp.append(item)
    c6 = temp[0]

    return c6


def print_wind(hwnd):
    print(hwnd, win32gui.GetWindowText(hwnd), ' ', win32gui.GetClassName(hwnd))


def keybord_input(key):
    win32api.keybd_event(VK_CODE[key], 0, 0, 0)
    win32api.keybd_event(VK_CODE[key], 0, win32con.KEYEVENTF_KEYUP, 0)


def select_all_copy():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('a')
    keybord_input('c')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)


def select_all():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('a')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)


def keybord_paste():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('v')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)
    # keybord_input('enter')


def get_input_pos():

    # 点击测试窗按钮,定位输入框
    win32api.SetCursorPos([int(GetSystemMetrics(0)-15),
                           int(GetSystemMetrics(1) - bot_height - 160)])
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    win32api.SetCursorPos([int(GetSystemMetrics(
        0)-random.randint(100, 200)), int(GetSystemMetrics(1) - bot_height - 20)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.1)


def check_answer_match():
    global bot_height
    # 全选答案
    time.sleep(0.6)
    win32api.SetCursorPos([int(GetSystemMetrics(0)-100),
                           int(GetSystemMetrics(1) - bot_height - 80)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    select_all_copy()
    time.sleep(0.6)
    answer = pyperclip.paste()
    answer_list = answer.split('\n')
    answer_list = answer_list[::-1]
    target_ans = ''
    for i in range(0, len(answer_list)):
        if (answer_list[i].find('命中问法') > -1) or (answer_list[i].find('命中关键字') > -1):
            if (answer_list[i-1].find('高频') > -1) or (answer_list[i-1].find('暂无') > -1):
                target_ans = '暂无'
            else:
                target_ans += answer_list[i-1].replace('自定义', '')
            break
        if answer_list[i].find('命中尺码表') > -1:
            target_ans = '暂无'
            break
    # 定位输入框
    win32api.SetCursorPos([int(GetSystemMetrics(
        0)-random.randint(100, 200)), int(GetSystemMetrics(1) - bot_height - 20)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    return target_ans


def turn2morequs():
    global bot_height
    keybord_input('F5')
    time.sleep(4)
    # 关闭弹出页面
    for wind in get_top_windows():
        if win32gui.GetClassName(wind) == 'Shell_TrayWnd':
            bot_height = win32gui.GetWindowRect(
                wind)[3] - win32gui.GetWindowRect(wind)[1]
    w_height = GetSystemMetrics(1)-100-bot_height
    height = int(w_height/2)-140+100
    width = int((GetSystemMetrics(0))/2)+285
    win32api.SetCursorPos([width, height])
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    # 截取屏幕分析按键位置
    qzm_tools.ImgFinder.window_capture()
    # 点击更多问题，方便将来导入
    win32api.SetCursorPos(get_element('more_qus.png'))
    # win32api.SetCursorPos([880, 230+alert_height])

    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.8)

def start_main():
    # 初始化到更多问题界面
    global hwndList
    global parent_hwnd
    global hwnd
    global alert_height

    try:
        hwnd = find_target_hwnd(parent_hwnd)
        title = win32gui.GetWindowText(hwnd)
        clsname = win32gui.GetClassName(hwnd)
        print('获取到工作台窗口')
        print('窗口句柄: ' + str(hwnd))
        print('窗口标题: ' + title)
        print('窗口类名: ' + clsname + '\n')
    except:
        print('未能获取到工作台窗口')
        return

    # 置顶对话框
    win32gui.BringWindowToTop(hwnd)
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    #win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, GetSystemMetrics(0),
    #                      GetSystemMetrics(1) - bot_height, win32con.SWP_SHOWWINDOW)

    # 获取网址编辑框并输入
    # edit_hwnd = get_edit_hwnd(hwnd)
    # edit_pos_left = win32gui.GetWindowRect(edit_hwnd)[0] + 5
    # edit_pos_right = win32gui.GetWindowRect(edit_hwnd)[1] + 5
    # win32api.SetCursorPos([edit_pos_left, edit_pos_right])
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    # pyperclip.copy('https://alphax.taobao.com/commercialize.html?appkey=23542470#/knowledge/knowledgeSolution')
    # #pyperclip.copy('店小蜜')
    # keybord_paste()
    # time.sleep(1)
    # keybord_input('enter')

    # time.sleep(4)
    # 设置窗口焦点
    win32gui.SetForegroundWindow(hwnd)

    # 跳转到更多问题界面
    turn2morequs()

def input_question(ques):
    pyperclip.copy(ques)
    select_all()
    keybord_paste()
    keybord_input('enter')
    time.sleep(0.2)


def clear_excel(num):
    pythoncom.CoInitialize()
    current_path = os.path.abspath(__file__)
    xapp = win32com.client.Dispatch('Excel.Application')
    xapp.DisplayAlerts = 0
    xapp.visible = 0
    xapp.ScreenUpdating = 0  # 关闭屏幕刷新
    xfile = xapp.Workbooks.Open(os.path.dirname(
        current_path) + '\\结果表格\\整理完毕问法.xlsx')
    sheet = xfile.Worksheets('Sheet1')
    for i in range(0, num):
        sheet.Rows(1).Delete()
    xfile.Close(SaveChanges=1)
    xapp.Quit()
    print('delete done')
    pythoncom.CoUninitialize()


def add_custom_answer(ques):
    global alert_height
    # 关闭测试窗
    win32api.SetCursorPos([int(GetSystemMetrics(0)-15), 110])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.8)
    # 点击添加按钮
    win32api.SetCursorPos(get_element('add_custom.png'))
    # win32api.SetCursorPos([GetSystemMetrics(0)-120, 480 + alert_height])

    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    # 输入内容
    win32api.SetCursorPos(
        [int(GetSystemMetrics(0)/2), int(GetSystemMetrics(1)/2-49)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    pyperclip.copy(ques)
    time.sleep(0.2)
    keybord_paste()
    time.sleep(0.3)
    # 点击保存按钮
    win32api.SetCursorPos(get_element('save_custom_qus.png'))
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.8)
    # 点击强行添加按钮
    try:
        # 可能没有强行确认按钮
        time.sleep(0.2)
        win32api.SetCursorPos(get_element('force_confirm.png'))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(0.5)
    except Exception as e:
        print(e)
        traceback.print_exc()
        print('没有强行确定按键')

    # 为避免提示重复问题，页面点两下
    win32api.SetCursorPos([int(GetSystemMetrics(
        0) - random.randint(100, 200)), int(GetSystemMetrics(1) - bot_height - 20)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)

    # 图二，自定义问题后
    # 点击添加答案
    win32api.SetCursorPos(get_element('add_ans.png'))
    # win32api.SetCursorPos([int(GetSystemMetrics(0)/2), int(GetSystemMetrics(1)/2+130)+alert_height])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)

    # 图三, 出现问法槽，点击问法槽并粘贴
    win32api.SetCursorPos(
        [int(GetSystemMetrics(0) / 2), int(GetSystemMetrics(1) / 2 - 150)])
    time.sleep(0.2)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    time.sleep(0.3)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.3)
    keybord_paste()
    try:
        # 按答案编辑可能没有富文本编辑器
        win32api.SetCursorPos(get_element('ans_edit_confirm.png'))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(0.2)
    except Exception:
        print('没有富文本编辑框')

    # 错误就跳过
    #
    # try:
        # if find_ans_input() != None:
    # win32api.PostMessage(ans_hwnd, win32con.WM_PASTE, 0, 0)
    #time.sleep(0.3)
    # win32api.PostMessage(confirm_hwnd, win32con.BM_CLICK, 0, 0)
        # else:
            # print('未能定位到编辑框')
    # except:
    #     print('添加失败')
    time.sleep(0.3)
    # 点击保存
    win32api.SetCursorPos(
        [int(GetSystemMetrics(0)/2 + 400), int(GetSystemMetrics(1)/2+350)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(1)
    # 点击测试窗按钮,定位输入框
    win32api.SetCursorPos([int(GetSystemMetrics(0)-15),
                           int(GetSystemMetrics(1) - bot_height - 160)])
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)
    win32api.SetCursorPos([int(GetSystemMetrics(
        0)-random.randint(100, 200)), int(GetSystemMetrics(1) - bot_height - 20)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.1)


def end_main():
    global hwnd
    # 取消置顶对话框
    time.sleep(0.1)
    win32gui.SetWindowPos(hwnd, win32con.HWND_BOTTOM, 0, 0, GetSystemMetrics(0), GetSystemMetrics(
        1) - bot_height, win32con.SWP_NOACTIVATE | win32con.SWP_NOOWNERZORDER | win32con.SWP_SHOWWINDOW)


def main():
    # 对表格处理读取数据
    print(os.listdir())
    unmatched_ques_list = []
    get_excel = load_workbook(os.path.join('src', 'code', 'secondTest','结果表格', '整理完毕问法.xlsx'))
    get_wb = get_excel.active
    for i in range(1, get_wb.max_row+1):
        unmatched_ques_list.append(get_wb.cell(row=i, column=1).value)

    print('读取到识别错误的问题数: ' + str(len(unmatched_ques_list)))
    start_time = time.time()

    # 测试问题的前页面的初始化
    start_main()
    # 开始测试问题并且处理
    get_input_pos()
    result_list = []
    i = 0
    for ques in unmatched_ques_list:
        i += 1
        # str(time.time())
        input_question(ques)
        # 分析是否匹配
        target_ans = check_answer_match()
        if target_ans == '':
            target_ans = check_answer_match()
        if target_ans == '暂无':
            #             # 添加自定义,并保存
            temp = [1, ques]
            result_list.append(temp)

            add_custom_answer(ques)
            wb = openpyxl.load_workbook('结果表格\添加自定义结果.xlsx')
            sheet = wb.active
            m_row = sheet.max_row
            _ = sheet.cell(row=m_row+1, column=1, value=ques)
            _ = sheet.cell(row=m_row+1, column=2, value=ques)
            wb.save(filename='结果表格\添加自定义结果.xlsx')
            print('问题: ' + ques + ' 答案: ' + ques)
        else:
            # 记录匹配的问题
            wb = openpyxl.load_workbook('结果表格\添加自定义结果.xlsx')
            sheet = wb.active
            m_row = sheet.max_row
            _ = sheet.cell(row=m_row+1, column=1, value=ques)
            _ = sheet.cell(row=m_row+1, column=2, value=target_ans.strip())
            wb.save(filename='结果表格\添加自定义结果.xlsx')
            print('问题: ' + ques + ' 答案: ' + target_ans.strip())
        if i == 70:
            clear_excel(70)
            i = 0
            turn2morequs()
            time.sleep(8)
            get_input_pos()

    end_main()


if __name__ == '__main__':
        main()
        traceback.print_exc()