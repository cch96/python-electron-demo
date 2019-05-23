# coding:utf-8
import time
import traceback
import re
import os

import win32api
import win32con
import win32gui
from win32api import GetSystemMetrics
from openpyxl import load_workbook
import pyperclip

import qzm_tools

POS = win32api.GetCursorPos()

VK_CODE = {
    'a': 0x41,
    'ctrl': 0x11,
    'c': 0x43,
    'v': 0x56,
    'enter': 0x0D,
    'down_arrow': 0x28,
    'F5': 0x74
}

# 顶部提示框高度
top_bar_height = 0
# 底部提示高度
bot_height = 0
time_active_list = []

elements = {}
# [('问法', ['答案1'， '答案2'...]), ...]
import_finish_list = []

# 自定义异常类
class NotFindQuestionError(Exception):
    """没有找到问题,把问题记录到excel中，继续下一个问题 """
    pass


class NeedArtificalAddError(Exception):
    """该问题的答案有人工话术按钮，需要人工对数据进行处理"""
    pass


class MultiFindQuestionError(Exception):
    """查出多条相似答案"""
    pass


class FindPicError(Exception):
    """无法找到上传图片"""
    pass


def error_handle_find_pic(question, answer):
    mouse_click(*get_element('ans_close.png', index=-1), delay=0.2)
    the_que = [question, [answer], '找不到图片']
    save_to_excel('结果表格/未导入答案.xlsx', the_que, max_col=6)


def error_handle_not_find_ques(single_qa_list):
    single_qa_list.append('没找到问题')
    save_to_excel('结果表格/未导入答案.xlsx', single_qa_list, max_col=6)


def error_handle_need_artifical_add(single_qa_list):
    single_qa_list.append('人工话术')
    keyboard_input('F5')
    wait_text('恢复默认', '更多问题')
    w_height = GetSystemMetrics(1)-100-bot_height
    height = int(w_height/2)-140+100
    width = int((GetSystemMetrics(0))/2)+285
    mouse_click(width, height, delay=0.3)


def error_handle_multi_find(single_qa_list):
    keyboard_input('F5')
    wait_text('恢复默认', '更多问题')
    w_height = GetSystemMetrics(1)-100-bot_height
    height = int(w_height/2)-140+100
    width = int((GetSystemMetrics(0))/2)+285
    mouse_click(width, height, delay=0.3)
    single_qa_list.append('找到多条问题')
    save_to_excel('结果表格/未导入答案.xlsx', single_qa_list, max_col=6)


def get_element(element_name, dirction=1, index=0):
    #  返回缓存的元素位置，若没有则截图对比获取位置
    global elements
    if element_name not in elements:
        qzm_tools.ImgFinder.window_capture()
        time.sleep(1.5)
        btnfinder =  qzm_tools.ImgFinder(element_name)
        elements[element_name] = btnfinder.get_element(dirction, index)
    print(element_name, elements[element_name])
    return elements[element_name]


def wait_text(*args, **kwargs):
    text_list = args
    result = [0] * len(text_list)
    reverse = kwargs.get('reverse', False)
    print("WAIT %s"  % ','.join(text_list) +  ' reverse:' + str(reverse))
    delay = kwargs.get('delay', 10)
    while keyboard_select_all(delay):
        content = pyperclip.paste()
        for i,text in enumerate(text_list):
            temp = content.find(text)
            if reverse:
                result[i] = 1  if temp < 0 else 0
            else:
                result[i] = 1  if temp > 0 else 0
        if all(result):
            break


def keyboard_select_all(delay=0.2):
    win32api.keybd_event(VK_CODE['ctrl'], 1, 0, 0)
    keyboard_input('a')
    keyboard_input('c')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(delay)
    return True


def keyboard_input(key):
    win32api.keybd_event(VK_CODE[key], 0, 0, 0)
    win32api.keybd_event(VK_CODE[key], 0, win32con.KEYEVENTF_KEYUP, 0)


def keyboard_paste():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keyboard_input('v')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)


def get_edit_hwnd(hwnd):
    c1 = None
    c1_list = []
    temp = []
    win32gui.EnumChildWindows(
        hwnd, lambda hWnd, param: param.append(hWnd), c1_list)
    for item in c1_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c2 = temp[1]
    temp = []
    c2_list = []

    win32gui.EnumChildWindows(
        c2, lambda hWnd, param: param.append(hWnd), c2_list)
    for item in c2_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c3 = temp[1]
    temp = []
    c3_list = []

    win32gui.EnumChildWindows(
        c3, lambda hWnd, param: param.append(hWnd), c3_list)
    for item in c3_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c4 = temp[0]
    temp = []
    c4_list = []

    win32gui.EnumChildWindows(
        c4, lambda hWnd, param: param.append(hWnd), c4_list)
    for item in c4_list:
        if win32gui.GetClassName(item) == 'StandardWindow':
            temp.append(item)
    c5 = temp[0]
    temp = []
    c5_list = []
    win32gui.EnumChildWindows(
        c5, lambda hWnd, param: param.append(hWnd), c5_list)
    for item in c5_list:
        if win32gui.GetClassName(item) == 'EditComponent':
            temp.append(item)
    c6 = temp[0]

    return c6


def get_top_windows():
    windList = []
    win32gui.EnumWindows(lambda hWnd, param: param.append(hWnd), windList)
    return windList


def set_wind_pos():
    global hwnd
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('- 工作台') > -1:
            hwnd = wind

    # 置顶并且获得焦点
    win32gui.BringWindowToTop(hwnd)
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(hwnd)


def clear_wind_pos():
    win32gui.SetWindowPos(hwnd, win32con.HWND_BOTTOM, GetSystemMetrics(0)-1280, 0, 1280, GetSystemMetrics(
        1)-60, win32con.SWP_NOACTIVATE | win32con.SWP_NOOWNERZORDER | win32con.SWP_SHOWWINDOW)


def mouse_click(pos1, po2, delay=0.0):
    win32api.SetCursorPos([pos1, po2])
    time.sleep(delay)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)


def get_time_active():
    mouse_click(GetSystemMetrics(0)-550, 450)
    time.sleep(0.1)
    keyboard_select_all()
    content = pyperclip.paste()
    # 取消全选
    mouse_click(GetSystemMetrics(0)-550, 450)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    content = content[0:content.find('测试窗')]
    content_list = content.split('\n')
    for i in range(0, len(content_list)):
        content_list[i] = content_list[i].strip()
        content_list[i] = content_list[i].replace('\r', '')
        content_list[i] = content_list[i].replace('\n', '')
        content_list[i] = content_list[i].replace('\t', '')
        if content_list[i] == '生效中' or content_list[i] == '未生效' or content_list[i] == '已失效':
            time_active_list.append(content_list[i-1])


def save_to_excel(file_path, the_que, max_col):
    wb = load_workbook(file_path)
    sheet = wb.active
    insert_row = sheet.max_row + 1

    ans_number = len(the_que[1])
    # 保存问题
    sheet.cell(row=insert_row, column=1, value=the_que[0])
    # 将问法根据答案的行数, 合并单元格
    if ans_number > 0:
        sheet.merge_cells(start_row=insert_row, start_column=1, end_row=insert_row+ans_number-1, end_column=1)

    # 保存答案及其他
    for no, ans in enumerate(the_que[1]):
        if ans:
            sheet.cell(row=no+insert_row, column=2, value=ans[0])  # ans
            sheet.cell(row=no+insert_row, column=3, value=ans[1])  # pic
            sheet.cell(row=no+insert_row, column=4, value='\n'.join(ans[2]))  # id_list
            sheet.cell(row=no+insert_row, column=5, value=ans[3])  # tac

    #  max_col为6则是未导入答案，表中需要多一个字段error_type
    if max_col == 6:
        error_import_type = the_que[2]
        sheet.cell(row=insert_row, column=6, value=error_import_type)
        # 将根据答案的行数, 合并error_type单元格
        if ans_number > 0:
            sheet.merge_cells(start_row=insert_row, start_column=6, end_row=insert_row+ans_number-1, end_column=6)

    # 准备下一次问题插入的行数
    insert_row += ans_number
    # 保存
    wb.save(filename = file_path)


def clear_excel_rows(file_path, offset, limit):
    wb = load_workbook(file_path)
    sheet = wb.active
    sheet.delete_rows(offset, limit)
    wb.save(filename = file_path)
    print('delete done')


def finish_save_excel(import_finish_list):
    for import_finish in import_finish_list:
        save_to_excel('结果表格/已导入答案.xlsx', import_finish, max_col=5)


def input_search(ques):
    mouse_click(*get_element('search_text.png'), delay=0.5)
    time.sleep(2)
    keyboard_select_all()
    pyperclip.copy(ques)
    keyboard_paste()
    keyboard_input('enter')

def step_search_ques(ques):
    # 搜索框搜索问题
    input_search(ques)
    print('搜索中……')
    mouse_click(GetSystemMetrics(0)-570, 250, 0.2)
    time.sleep(1)

    # 搜索问题, 等待搜索问题的结果
    while True:
        qzm_tools.ImgFinder.window_capture()
        try:
            # 有答案情况
            qzm_tools.ImgFinder('ans_delete_flag.png').get_element()
            break
        except qzm_tools.FindImgError:
            pass

        try:
            # 无答案情况
            qzm_tools.ImgFinder('add_ans.png').get_all_element()
            break
        except qzm_tools.FindImgError:
            pass

        try:
            # 没有找到问题情况
            qzm_tools.ImgFinder('not_find_flag.png').get_element()
            raise NotFindQuestionError('没有找到问题: %s' % ques)
        except qzm_tools.FindImgError:
            pass
    #  通过文字排除情况，避免图片识别不到的情况
    keyboard_select_all()
    content = pyperclip.paste()
    if content.count('恢复默认') > 1:
        raise MultiFindQuestionError('找到了多条问题')
    mouse_click(GetSystemMetrics(0)-670, 250)


def step_click_add_ans(add_ans_position):
    mouse_click(*add_ans_position)
    time.sleep(0.5)


def check_step_click_ans_input():
    opened_ans_input = False
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('编辑框') > -1:
            opened_ans_input = True
            break
    return opened_ans_input


def mouse_wheel_down():
    win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL, 0, 0, -10000)
    print('下滑')

def step_click_ans_input():
    mouse_click(GetSystemMetrics(0)-910, 500)
    text_position = qzm_tools.until_get_element('text_flag.png', fun=mouse_wheel_down)
    click_position_x, click_position_y = text_position[0]-300, text_position[1]-80
    mouse_click(click_position_x, click_position_y, delay=0.5)
    time.sleep(0.5)

def get_input_hwnd():
    par_hwnd = None
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('编辑框') > -1:
            par_hwnd = wind
            break
    return par_hwnd


def step_input_ans(ans):
    pyperclip.copy(ans)
    keyboard_paste()
    time.sleep(0.2)
    try:
        # 按答案编辑可能没有富文本编辑器
        win32api.SetCursorPos(get_element('ans_edit_confirm.png'))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(0.2)
    except Exception:
        print('没有富文本编辑框')
    time.sleep(0.3)


def step_click_pic_input():
    mouse_click(*get_element('add_pic.png'), delay=0.2)
    while not qzm_tools.until_get_element('upload_flag.png', attempts=2):
        mouse_click(*get_element('add_pic.png'))


def get_upload_hwnd():
    par_hwnd = None
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('打开') > -1:
            par_hwnd = wind
            break
    return par_hwnd


def deal_upload_file(pic):
    print('上传附件')
    pic_hwnd = None
    pic_confirm = None
    pic_cancel = None
    par_hwnd = get_upload_hwnd()
    no_pic = False
    c1 = []
    c1_temp = None
    search_box = None
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c1)
    for cwind in c1:
        if win32gui.GetClassName(cwind).find('ComboBoxEx32') > -1:
            c1_temp = cwind
        if win32gui.GetClassName(cwind) == 'SearchEditBoxWrapperClass':
            search_box = cwind
        if win32gui.GetWindowText(cwind).find('打开') > -1:
            pic_confirm = cwind
        if win32gui.GetWindowText(cwind).find('取消') > -1:
            pic_cancel = cwind

    c2 = []
    c2_temp = None
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c2)
    for cwind in c2:
        if win32gui.GetClassName(cwind).find('ComboBoxEx') > -1:
            c2_temp = cwind

    c3 = []
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c3)
    for cwind in c3:
        if win32gui.GetClassName(cwind).find('Edit') > -1:
            pic_hwnd = cwind
            break

    # 输入图片
    file_name = str(pic)
    win32gui.SetForegroundWindow(search_box)
    if pic_hwnd != None:
        if file_name.endswith('.txt'):
            pyperclip.copy(file_name)
            win32api.PostMessage(pic_hwnd, win32con.WM_PASTE, 0, 0)
            time.sleep(0.2)
            win32api.PostMessage(pic_confirm, win32con.BM_CLICK, 0, 0)
            time.sleep(0.2)
        # 除了txt文件其他都考虑为图片文件
        else:
            for file in os.listdir('附件'):
                if file.split('.')[0] == file_name and os.path.isfile(os.path.join('附件', file)):
                    pyperclip.copy(file)
                    win32api.PostMessage(pic_hwnd, win32con.WM_PASTE, 0, 0)
                    time.sleep(0.2)
                    win32api.PostMessage(pic_confirm, win32con.BM_CLICK, 0, 0)
                    time.sleep(0.2)
                    break
            else:
                # 三种图片都没有则取消上传图片
                win32api.PostMessage(pic_cancel, win32con.BM_CLICK, 0, 0)
                raise FindPicError('找不到图片')
    return qzm_tools.until_dis_element('upload_flag.png')


def step_input_pic(pic):
    while not deal_upload_file(pic):
        pass
    return True


def step_open_time_input(times):
    global time_active_list
    mouse_click(*get_element('related_time.png'))
    #win32api.SetCursorPos(get_element('related_time.png'))
    time.sleep(0.5)
    print(time_active_list, times)
    pos = time_active_list.index(times)+1
    if pos < 4:
        mouse_click(GetSystemMetrics(0) - 1150, 710+pos*70)
    else:
        win32api.SetCursorPos([GetSystemMetrics(0) - 1150, 910])
        for i in range(0, pos-3):
            keyboard_input('down_arrow')
            time.sleep(0.05)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.5)


def step_click_spe_item():
    mouse_click(*qzm_tools.until_get_element('designated_commodity.png', fun=mouse_wheel_down))
    while not qzm_tools.until_get_element('import_commodity.png', attempts=5):
        mouse_click(*qzm_tools.until_get_element('designated_commodity.png', fun=mouse_wheel_down))


def check_step_click_upload_txt():
    pic_hwnd = None
    pic_confirm = None
    pic_cancel = None
    par_hwnd = None
    has_hwnd = False
    for wind in get_top_windows():
        if win32gui.GetWindowText(wind).find('打开') > -1:
            par_hwnd = wind
            break
    c1 = []
    c1_temp = None
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c1)
    for cwind in c1:
        if win32gui.GetClassName(cwind).find('ComboBoxEx32') > -1:
            c1_temp = cwind
        if win32gui.GetWindowText(cwind).find('打开') > -1:
            pic_confirm = cwind
        if win32gui.GetWindowText(cwind).find('取消') > -1:
            pic_cancel = cwind

    c2 = []
    c2_temp = None
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c2)
    for cwind in c2:
        if win32gui.GetClassName(cwind).find('ComboBoxEx') > -1:
            c2_temp = cwind

    c3 = []
    win32gui.EnumChildWindows(par_hwnd, lambda hWnd,
                              param: param.append(hWnd), c3)
    for cwind in c3:
        if win32gui.GetClassName(cwind).find('Edit') > -1:
            pic_hwnd = cwind
            break
    if pic_hwnd != None:
        has_hwnd = True

    return has_hwnd


def step_click_upload_txt():
    mouse_click(*get_element('import_commodity.png'), delay=0.1)
    mouse_click(*qzm_tools.until_get_element('upload_text.png'), delay=0.1)
    while not qzm_tools.until_get_element('upload_flag.png'):
        mouse_click(*qzm_tools.until_get_element('upload_text.png'), delay=0.1)


def step_upload_txt():
    while not deal_upload_file('商品id.txt'):
        pass
    return True


def step_check_result():
    qzm_tools.until_get_element('import_finish_flag.png')
    mouse_click(*qzm_tools.until_get_element('import_finish.png'), delay=0.1)
    # 等待答案框消失
    qzm_tools.until_get_element('import_dis_flag.png')
    time.sleep(0.3)


def step_save_ans():
    # 点击保存
    mouse_click(*get_element('save.png'), delay=0.2)
    qzm_tools.until_dis_element('ans_flag.png')


def time_wait_add_ans(before_average_gray):
    # 等待页面问题，即mask消失，渲染完成
    qzm_tools.until_stable_window(before_average_gray)


def get_first_ques(ques):
    ques_list = []
    ques = str(ques)
    print(ques)
    ques.strip()
    ques.strip('/')
    ques = ques.replace('/', '\n')
    temp = ques.split('\n')
    for item in temp:
        item = item.replace('\r', '')
        item = item.replace('\n', '')
        item = item.replace('\t', '')
        if item != '':
            ques_list.append(item)
    # 只返回第一个
    return ques_list[0]


def get_id_list(id_str):
    id_str = str(id_str)
    temp = id_str.split('\n')
    id_list = []
    # print('get_id_list', id_str, temp)
    if id_str.find('全部商品') > -1:
        id_list.append('全部商品')
    else:
        for item in temp:
            item = re.sub('\D', '', item)
            # item = item+'<br>'
            # print('get_id_list', 'item', item)
            id_list.append(item)

    return id_list


def get_format_string(un_str):
    f_str = un_str
    if un_str != None:
        f_str = str(un_str).strip()
    return f_str


def get_excel_content():
    excel = load_workbook('结果表格/QA表.xlsx')
    sheet = excel.active
    m_row = sheet.max_row


    # qa_list   single_qa_list  q_lis    a_list  single_ans     ans         pic   id_list   tac
    #    [          [     [问题合集] ,   [       [           答案内容，   pic, [ 商品id ], tac  ]  ] ] ]

    qa_list = []

    start = 2
    for i in range(3, m_row+1):
        if sheet.cell(row=i, column=1).value != None:
            single_qa_list = []
            single_qa_list.append(
                get_first_ques(sheet.cell(row=start, column=1).value))  # q_list
            a_list = []  # a_list
            for a in range(start, i):
                single_ans = []
                single_ans.append(get_format_string(
                    sheet.cell(row=a, column=2).value))  # ans
                single_ans.append(get_format_string(
                    sheet.cell(row=a, column=3).value))  # pic
                single_ans.append(get_id_list(
                    sheet.cell(row=a, column=4).value))  # id_list
                single_ans.append(get_format_string(
                    sheet.cell(row=a, column=5).value))  # tac
                # print('get_id_list', 'single_ans', single_ans)
                a_list.append(single_ans)
                # print('get_id_list', 'a_list', a_list)
            single_qa_list.append(a_list)
            qa_list.append(single_qa_list)
            start = i
            if i == m_row:
                single_qa_list= []
                single_qa_list.append(sheet.cell(
                    row=start, column=1).value)  # type
                single_qa_list.append(
                    get_first_ques(sheet.cell(row=start, column=2).value))  # q_list
                a_list = []  # a_list
                for a in range(start, i):
                    single_ans = []
                    single_ans.append(get_format_string(
                        sheet.cell(row=a, column=3).value))  # ans
                    single_ans.append(get_format_string(
                        sheet.cell(row=a, column=4).value))  # pic
                    single_ans.append(get_id_list(
                        sheet.cell(row=a, column=5).value))  # id_list
                    single_ans.append(get_format_string(
                        sheet.cell(row=a, column=6).value))  # tac
                    a_list.append(single_ans)
                single_qa_list.append(a_list)
                qa_list.append(single_qa_list)
    print(qa_list)
    print('统计到问法总数:', len(qa_list))
    return qa_list


def step_check_ans_input():
    if not qzm_tools.until_get_element('ans_flag.png', attempts=10):
        return False
    # 不需要转人工话术
    # finder = qzm_tools.ImgFinder('artificial_flag.png')
    # try:
    #     finder.get_element()
    # except qzm_tools.FindImgError:
    #     pass
    # else:
    #     raise NeedArtificalAddError('需要转人工')
    return True


def step_see_add_button():
    global elements
    # 将鼠标移动于不会有弹出文字的地方
    win32api.SetCursorPos([GetSystemMetrics(0) - 240, 360])
    add_ans_position = qzm_tools.until_get_element('add_ans.png', index=-1, fun=mouse_wheel_down)
    elements['totop.png'] = get_element('totop.png', index=-1)
    before_average_gray =  qzm_tools.get_last_screen_gray()
    return (add_ans_position, before_average_gray)


def add_official_single_ans(ans, pic, id_list, tac):
    # print('add_official_single_ans',ques,ans,pic,id_list,tac)

    f = open('附件/商品id.txt', 'w')
    for item in id_list:
        f.write(item+'\n')
    f.close()

    # step0
    add_ans_position, before_average_gray = step_see_add_button()

    # step1
    step_click_add_ans(add_ans_position)

    # step2
    while not step_check_ans_input():
        add_ans_position, before_average_gray = step_see_add_button()
        step_click_add_ans(add_ans_position)

    # step3
    step_click_ans_input()

    # step4
    step_input_ans(ans)

    # step5 下拉到底
    mouse_click(GetSystemMetrics(0)-910, 280)
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL, 0, 0, -10000)
    time.sleep(1)

    if pic != None:
        # step6
        step_click_pic_input()

        # step7
        step_input_pic(pic)


    if tac != '永久生效' and bool(tac) != False:
        # step11
        step_open_time_input(tac)
    if id_list[0] != '全部商品' and bool(id_list[0]) != False:
        # step9
        step_click_spe_item()

        # step10
        step_click_upload_txt()

        # step11
        step_upload_txt()

        # step12
        step_check_result()

    # step13
    step_save_ans()

    # step14
    time_wait_add_ans(before_average_gray)
    return True


def keybord_input(key):
    win32api.keybd_event(VK_CODE[key], 0, 0, 0)
    win32api.keybd_event(VK_CODE[key], 0, win32con.KEYEVENTF_KEYUP, 0)


def add_all_ans(single_qa_list):
    all_ans = single_qa_list[1]
    for i, single_ans in enumerate(all_ans):
        print('    添加答案', i+1, '/', len(all_ans))
        try:
            add_official_single_ans(single_ans[0],
                              single_ans[1], single_ans[2], single_ans[3])
        except FindPicError:
            question = single_qa_list[0]
            error_handle_find_pic(question, single_ans)
            all_ans.pop(i)
            continue
    return all_ans


def main():
    global import_finish_list
    global bot_height
    global elements
    # start
    set_wind_pos()
    time.sleep(0.5)
    for wind in get_top_windows():
        if win32gui.GetClassName(wind) == 'Shell_TrayWnd':
            bot_height = win32gui.GetWindowRect(
                wind)[3] - win32gui.GetWindowRect(wind)[1]
    # 进入时效页
    # website_edit_input('https://alphax.taobao.com/commercialize.html?appkey=23542470#/knowledge/knowledgeActivity')
    mouse_click(*get_element('knowledge_activity.png'))
    time.sleep(2)
    mouse_click(GetSystemMetrics(0)-350, 420)
    time.sleep(0.3)

    # 获取时效
    get_time_active()
    print(time_active_list)

    # 进入问法页
    # website_edit_input('https://alphax.taobao.com/commercialize.html?appkey=23542470#/knowledge/knowledgeSolution')
    time.sleep(0.5)
    mouse_click(*get_element('knowledge_solution.png'))
    time.sleep(2)
    mouse_click(GetSystemMetrics(0)-350, 420)
    time.sleep(0.3)
    # 关闭行业高频问题推荐框，关闭这个框之后回答问题不会出现推荐框
    try:
        mouse_click(*get_element('recommend_cross.png'), delay=0.2)
        time.sleep(0.5)
    except Exception as e:
        print('没有推荐框')

    # qa_list   class_list   q_list     a_list  single_ans   ans     pic   id_list     tac
    #    [        [        [问题合集] ,   [        [       答案内容，pic, [ 商品id ], tac  ]  ] ] ]

    qa_list = get_excel_content()

    # 记录excel取数据到第几行，删除用
    get_excel_row = 2
    for i, single_qa_list in enumerate(qa_list):
        try:
            step_search_ques(single_qa_list[0])  # 搜索问题
            import_finish_ans = add_all_ans(single_qa_list)  # 添加答案
        except NotFindQuestionError as e:
            print(e)
            error_handle_not_find_ques(single_qa_list)
            continue
        except NeedArtificalAddError as e:
            print(e)
            error_handle_need_artifical_add(single_qa_list)
            continue
        except MultiFindQuestionError as e:
            print(e)
            error_handle_multi_find(single_qa_list)
            continue
        else:
            import_finish_list.append([single_qa_list[0], import_finish_ans])
            # 到达页面顶部
            # win32api.mouse_event(win32con.MOUSEEVENTF_WHEEL, 0, 0, 10000)
            mouse_click(*get_element('totop.png'))
            qzm_tools.until_get_element('top_flag.png')

        # 导入后或处理错误后
        # get_excel_row += len(single_qa_list[1])
        #  每十条问题保存进excel,删除原文件10个问题,
        #  或完成所有问题则保存（不包括完成最后一条标志结束记录)
        # if i %10 == 9 or i == len(qa_list)-2:
        #     save_to_excel('结果表格/已导入答案.xlsx', import_finish_list, max_col=5)
        #     import_finish_list = []
        #     save_to_excel('结果表格/未导入答案.xlsx', import_error_list, max_col=6)
        #     import_error_list = []
        #     clear_excel_rows('结果表格/QA表.xlsx', 2, get_excel_row-1)
        #     get_excel_row = 2

        get_excel_row += len(single_qa_list[1])
        if i % 10 == 9 or i == len(qa_list) - 2:
            finish_save_excel(import_finish_list)
            import_finish_list = []
            clear_excel_rows('结果表格/QA表.xlsx', 2, get_excel_row - 1)
            get_excel_row = 2

    # 关闭取消置顶窗口
    time.sleep(1)
    clear_wind_pos()


if __name__ == '__main__':
    try:
        main()
    except:
        traceback.print_exc()
    input()
