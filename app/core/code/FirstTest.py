# coding:utf-8
import win32api, win32con, win32gui
from win32api import GetSystemMetrics
import openpyxl
from openpyxl import load_workbook
import pyperclip
import time
import random
import os
import threading
import tkinter
from tkinter import *
import tkinter.messagebox
import win32com.client
import pythoncom

VK_CODE = {
    'a':0x41,
    'ctrl':0x11,
    'c':0x43,
    'v':0x56,
    'enter':0x0D,
    'F5':0x74
}

#初始ui内容
msg = None
initmsg = '--------------------------------------'
btn = None

root = tkinter.Tk()
root.geometry('200x500')
root.geometry('+0+0')

timestamp = None

chat_wnd = None
hwndList = None
parent_hwnd = None
bot_height = 0
final_result_string = ''
matched_num = 0
unmatched_num = 0

def analysis_answer(answer_list):
    global initmsg
    global msg
    global matched_num
    global unmatched_num
    matched_ans_list = []
    unmatched_ans_list = []
    excel_ans_list = []
    ques = ''
    match = ''
    answ = ''
    mzwf = 0
    xsda = 0
    fk = -1

    for i in range(0, len(answer_list)):
        if (answer_list[i].find('命中问法') > -1) or (answer_list[i].find('命中尺码表') > -1) or (answer_list[i].find('命中关键字') > -1):
            mzwf = i
            for a in range(fk+1, mzwf):
                ques += answer_list[a]
            ques = ques.replace('\r', '')
        if answer_list[i].find('显示答案') > -1:
            xsda = i
            for b in range(mzwf+1, xsda):
                match += answer_list[b]
            match = match.replace('\r', '')
        if answer_list[i].find('有误反馈给店小蜜') > -1:
            fk = i
            for c in range(xsda+1, fk):
                answ += answer_list[c]
            answ = answ.replace('\r', '')
            excel_ans_list.append([ques, match, answ])
            ques = ''
            match = ''
            answ = ''
    newtext = '\nall ans: ' + str(len(excel_aeds_list))
    initmsg += newtext
    msg['text'] = initmsg
    print('all ans: ' + str(len(excel_ans_list)))

    #删除已经处理的问题
    clear_excel(len(excel_ans_list))
    
    for ans in excel_ans_list:
        if ans[1] == '暂无':
            unmatched_ans_list.append(ans)
        else:
            matched_ans_list.append(ans)

    #检查进度
    wb1 = openpyxl.load_workbook('结果表格/未被官方识别问题.xlsx')
    m_row1 = wb1.active.max_row + len(unmatched_ans_list)
    wb1.save(filename = '结果表格/未被官方识别问题.xlsx')

    wb2 = openpyxl.load_workbook('结果表格/被官方识别问题.xlsx')
    m_row2 = wb2.active.max_row + len(matched_ans_list)
    wb2.save(filename = '结果表格/被官方识别问题.xlsx')

    wb3 = openpyxl.load_workbook('结果表格/所有问题集合.xlsx')
    m_row3 = wb3.active.max_row
    wb3.save(filename = '结果表格/所有问题集合.xlsx')
    #
    newtext = '\nmatched ans: ' + str(len(matched_ans_list)) + '\nunmatched ans: ' + str(len(unmatched_ans_list))
    newtext += '\n总计筛选出未被识别问题数: ' + str(m_row1)
    newtext += '\n总计筛选出被识别问题数: ' + str(m_row2)
    newtext += '\n总计剩余问题数: ' + str(m_row3)
    initmsg = '--------------------------------------'
    initmsg += newtext
    msg['text'] = initmsg
    print('matched ans: ' + str(len(matched_ans_list)))
    print('unmatched ans: ' + str(len(unmatched_ans_list)))
    save_to_excel(True, matched_ans_list)
    save_to_excel(False, unmatched_ans_list)
    matched_num += len(matched_ans_list)
    unmatched_num += len(unmatched_ans_list)

def save_to_excel(match, result_list):
    wb = openpyxl.load_workbook('结果表格/未被官方识别问题.xlsx')
    if match:
        wb = openpyxl.load_workbook('结果表格/被官方识别问题.xlsx')
    sheet = wb.active
    m_row = sheet.max_row

    for the_row in range(1, len(result_list)+1):
        if m_row == 1:
            m_row = 0
        for cell in range(1, 4):
            _ = sheet.cell(row=the_row+m_row, column=cell, value=result_list[the_row-1][cell-1])
    if match:
        wb.save(filename = '结果表格/被官方识别问题.xlsx')
    else:
        wb.save(filename = '结果表格/未被官方识别问题.xlsx')

def check_chat_window(hWnd):
    if not hWnd:
        return Falsse
    if '- 工作台' in win32gui.GetWindowText(hWnd):
        return True

def check_bot_window(hWnd):
    if not hWnd:
        return Falsse
    if 'MSTaskListWClass' in win32gui.GetClassName(hWnd):
        return True

def get_top_windows():
    windList = []
    win32gui.EnumWindows(lambda hWnd, param: param.append(hWnd), windList)
    return windList

def find_target_hwnd(parent_hwnd):
    child_list = []
    win32gui.EnumWindows(lambda hwnd, param: param.append(hwnd), child_list)
    for wind in child_list:
        if check_chat_window(wind):
            first_wind = wind
            break
    return first_wind

def get_edit_hwnd(hwnd):
    c1 = None
    c1_list = []
    temp = []
    win32gui.EnumChildWindows(hwnd, lambda hWnd, param: param.append(hWnd), c1_list)
    for item in c1_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c2 = temp[1]
    temp = []
    c2_list = []
    
    win32gui.EnumChildWindows(c2, lambda hWnd, param: param.append(hWnd), c2_list)
    for item in c2_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c3 = temp[1]
    temp = []
    c3_list = []
    
    win32gui.EnumChildWindows(c3, lambda hWnd, param: param.append(hWnd), c3_list)
    for item in c3_list:
        if win32gui.GetClassName(item) == 'StackPanel':
            temp.append(item)
    c4 = temp[0]
    temp = []
    c4_list = []
    
    win32gui.EnumChildWindows(c4, lambda hWnd, param: param.append(hWnd), c4_list)
    for item in c4_list:
        if win32gui.GetClassName(item) == 'StandardWindow':
            temp.append(item)
    c5 = temp[0]
    temp = []
    c5_list = []
    win32gui.EnumChildWindows(c5, lambda hWnd, param: param.append(hWnd), c5_list)
    for item in c5_list:
        if win32gui.GetClassName(item) == 'EditComponent':
            temp.append(item)
    c6 = temp[0]
    
    return c6
'''
def print_wind(hwnd):
    print(hwnd, win32gui.GetWindowText(hwnd), ' ', win32gui.GetClassName(hwnd))
'''
def keybord_input(key):
    win32api.keybd_event(VK_CODE[key], 0, 0, 0)
    win32api.keybd_event(VK_CODE[key], 0, win32con.KEYEVENTF_KEYUP, 0)

def select_all_copy():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('a')
    keybord_input('c')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)

def select_all():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('a')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)

def keybord_paste():
    win32api.keybd_event(VK_CODE['ctrl'], 0, 0, 0)
    keybord_input('v')
    win32api.keybd_event(VK_CODE['ctrl'], 0, win32con.KEYEVENTF_KEYUP, 0)
    #keybord_input('enter')

def get_input_pos():
    #关闭弹出页面
    global bot_height
    w_height = GetSystemMetrics(1)-100-bot_height
    height = int(w_height/2)-140+100
    width = int((GetSystemMetrics(0)-200)/2)+200+285
    win32api.SetCursorPos([width, height])
    time.sleep(0.5)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.5)
    #点击测试窗按钮,定位输入框
    win32api.SetCursorPos([int(GetSystemMetrics(0)-15), int(GetSystemMetrics(1) - bot_height - 160)])
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(1)
    win32api.SetCursorPos([int(GetSystemMetrics(0)-random.randint(100, 200)), int(GetSystemMetrics(1) - bot_height - 20)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    time.sleep(0.1)

def get_part_answer():
    global initmsg
    global msg
    global timestamp
    global final_result_string
    global bot_height
    newtext = '\n\n' + timestamp
    initmsg += newtext
    msg['text'] = initmsg
    print('\n' + timestamp)
    #全选答案
    win32api.SetCursorPos([int(GetSystemMetrics(0)-50), int(GetSystemMetrics(1) - bot_height - 80)])
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    select_all_copy()
    time.sleep(1)
    try:
        answer = pyperclip.paste()
        answer = answer[answer.index(timestamp):len(answer)]
        answer = answer[answer.index('蜜')+3:len(answer)-3]
        final_result_string += answer
    except:
        newtext = '\n分析问题答案时发生错误'
        initmsg += newtext
        msg['text'] = initmsg
        print('get error in get part answer')
        raise

def clear_excel(num):
    pythoncom.CoInitialize()
    current_path = os.path.abspath(__file__)
    xapp = win32com.client.Dispatch('Excel.Application')
    xfile = xapp.Workbooks.Open(os.path.dirname(current_path) + '\\结果表格\\所有问题集合.xlsx')
    sheet = xfile.Worksheets('Sheet1')
    for i in range(0, num):
        sheet.Rows(1).Delete()
    xfile.Close(SaveChanges=1)
    print('delete done')
    pythoncom.CoUninitialize()

def start_main():
    #获取对话框句柄
    global hwndList
    global parent_hwnd
    global hwnd
    global bot_height

    global msg
    global initmsg
    try:
        hwnd = find_target_hwnd(parent_hwnd)
        title = win32gui.GetWindowText(hwnd)
        clsname = win32gui.GetClassName(hwnd)
        print('获取到工作台窗口')
        print('窗口句柄: ' + str(hwnd))
        print('窗口标题: ' + title)
        print('窗口类名: ' + clsname + '\n')
        newtext = '\n获取到工作台窗口' + '\n窗口句柄: ' + str(hwnd) + '\n窗口标题: ' + title + '\n窗口类名: ' + clsname + '\n'
        initmsg += newtext
        msg['text'] = initmsg
    except:
        newtext = '未能获取到工作台窗口'
        initmsg += newtext
        msg['text'] = initmsg
        print('未能获取到工作台窗口')
        raise
        return

    for wind in get_top_windows():
        if win32gui.GetClassName(wind) == 'Shell_TrayWnd':
            bot_height = win32gui.GetWindowRect(wind)[3] - win32gui.GetWindowRect(wind)[1]

    #置顶对话框
    win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST
                              , 200
                              , 0, GetSystemMetrics(0)-200,GetSystemMetrics(1) - bot_height
                              , win32con.SWP_NOACTIVATE| win32con.SWP_NOOWNERZORDER|win32con.SWP_SHOWWINDOW)

    #获取网址编辑框并输入
    # edit_hwnd = get_edit_hwnd(hwnd)
    # edit_pos_left = win32gui.GetWindowRect(edit_hwnd)[0] + 5
    # edit_pos_right = win32gui.GetWindowRect(edit_hwnd)[1] + 5
    # win32api.SetCursorPos([edit_pos_left, edit_pos_right])
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
    # win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
    # pyperclip.copy('https://alphax.taobao.com/commercialize.html?appkey=23542470#/knowledge/knowledgeSolution')
    # keybord_paste()
    # time.sleep(1)
    # keybord_input('enter')

 

    keybord_input('F5')
    time.sleep(8)
    get_input_pos()

def input_question(ques):
    global initmsg
    global msg
    try:
        pyperclip.copy(ques)
        select_all()
        keybord_paste()
        keybord_input('enter')
        time.sleep(0.2)
    except:
        newtext = '\nerror ques: ' + str(ques)
        initmsg += newtext
        msg['text'] = initmsg
        print('error ques: ' + str(ques))
        raise

def end_main():
    global hwnd
    #取消置顶对话框
    time.sleep(0.1)
    win32gui.SetWindowPos(hwnd, win32con.HWND_BOTTOM
                              , 200
                              , 0, GetSystemMetrics(0)-200,GetSystemMetrics(1) - bot_height
                              , win32con.SWP_NOACTIVATE| win32con.SWP_NOOWNERZORDER|win32con.SWP_SHOWWINDOW)

def main_start_thread():
    t = threading.Thread(target=click_yes)
    t.start()

def click_yes():
    global initmsg
    global btn
    global msg
    if tkinter.messagebox.askquestion(message='在点击"是"以前请完成以下项目:\n1. 启动并登陆千牛工作台，不要最小化\n2. 聊天记录存为聊天记录.txt文件\n3. 尽量关闭可能弹出窗口的软件，比如杀毒软件') == 'yes':
        initmsg = '--------------------------------------'
        msg['text'] = initmsg
        btn['text'] = '运行中'
        btn['state'] = tkinter.DISABLED
        initmsg += '\n程序开始运行,请勿操作鼠标键盘'
        msg['text'] = initmsg
        time.sleep(1)
        start_main_run()
        btn['state'] = tkinter.NORMAL
        btn['text'] = '运行程序'

def start_main_run():
    try:
        main_run()
    except:
        print('出现错误，重新运行')
        start_main_run()

def main_run():
    #main
    global msg
    global initmsg
    global timestamp
    global final_result_string
    start_time = time.time()
    start_main()

    question_list = []
    excel = load_workbook('结果表格/所有问题集合.xlsx')
    table = excel.active
    rows = table.max_row
    for i in range(1, rows+1):
        question_list.append(table.cell(row=i, column=1).value)

    newtext = '\n总问题数: ' + str(len(question_list))
    initmsg += newtext
    msg['text'] = initmsg
    print('总问题数: ' + str(len(question_list)))

    timestamp = str(time.time())
    input_question(timestamp)
    i = 0
    for ques in question_list:
        #str(time.time())
        input_question(ques)
        i += 1
        if i == 100:
            #分析并获取结果
            get_part_answer()
            try:
                answer_list = final_result_string.split('\n')
                t = threading.Thread(target=analysis_answer, args=([answer_list]))
                t.start()
                time.sleep(2)
                #analysis_answer(answer_list)
                final_result_string = ''
            except: 
                newtext = '\n获取问题及答案时发生错误'
                initmsg += newtext
                msg['text'] = initmsg
                print('get error in analysis part answer')
            win32api.SetCursorPos([GetSystemMetrics(0)-430, GetSystemMetrics(1)-200])
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,0,0,0,0)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,0,0,0,0)
            keybord_input('F5')
            time.sleep(5)
            get_input_pos()
            timestamp = str(time.time())
            input_question(timestamp)
            i = 0

    time.sleep(1)

    #分析并获取结果
    try:
        get_part_answer()
        answer_list = final_result_string.split('\n')
        analysis_answer(answer_list)
        final_result_string = ''
    except:
        
        newtext = '\n分析问题答案时发生错误'
        initmsg += newtext
        msg['text'] = initmsg
        print('get error in analysis part answer')

    #所有问题都已存入excel
    
    newtext = '\n\n本次任务中:\n被官方识别的问题数: ' + str(matched_num)
    newtext += '\n未被官方识别的问题数目: ' + str(unmatched_num)
    initmsg += newtext
    msg['text'] = initmsg
    print('被官方识别的问题数: ' + str(matched_num))
    print('未被官方识别的问题数目: ' + str(unmatched_num))
    matched_num = 0
    unmatched_num = 0

    end_main()
    newtext = '\n\n总耗时(秒): ' + str(int(time.time()) - int(start_time))
    initmsg += newtext
    msg['text'] = initmsg
    print('\n总耗时(秒): ' + str(int(time.time()) - int(start_time)))


def main():
    lbl = Label(root, text='官方识别测试')
    lbl.pack()

    btn = Button(root, text = '运行程序', command=main_start_thread)
    btn.pack()

    msg = Message(root, text = initmsg, width=400)
    msg.config(bg='white')
    msg.pack()

    root.mainloop()


if __name__ == '__main__':
    main()


