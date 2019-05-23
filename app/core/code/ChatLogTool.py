# coding:utf-8
import openpyxl
from openpyxl import load_workbook
import re

zh_pattern = re.compile(u'[\u4e00-\u9fa5]+')

content = ''
content_list = []
store_name = ''

def contain_zh(word):
    global zh_pattern
    match = zh_pattern.search(word)

    return match

def check_ques(ques):
    valid_ques = True
    
    if not (ques.find('):') > -1) or (ques.find(store_name) == 0):
        valid_ques = False
    else:
        ques = ques[ques.find('):')+2: len(ques)]
        ques = ques.strip()
        ques = ques.replace('[语音]', '')
        ques = ques.replace('[文件]', '')
        ques = ques.replace('[图片]', '')
        ques = ques.replace('[卡片]', '')
        ques = ques.replace('[表情]', '')
        ques = ques.replace('[未知]', '')
        ques = ques.replace('[emoji]', '')
    if ques == '':
        valid_ques = False
    if len(ques) < 4 or len(ques) > 25:
        valid_ques = False
    if not contain_zh(ques):
        valid_ques = False
    num = 0
    for c in ques:
        try:
            c = int(c)
            num += 1
        except:
            num = num
        if num > 10:
            valid_ques = False
            break
        
    if valid_ques:
        valid_ques = ques
    return valid_ques
        
file = open('聊天记录.txt', 'rt', encoding ='gb18030', errors='ignore')
print('开始筛选聊天记录')
'''
for line in file:
    content += line
'''
content = file.read()
temp = content.split('\n')
store_name = temp[0].strip()
store_name = store_name[0:store_name.find(':')]
print('店铺名称: ' + store_name)

for ques in temp:
    result = check_ques(ques)
    if not result == False:
        content_list.append(result)

print('筛选出问题总数: ' + str(len(content_list)))
print('开始存入所有问题集合.xlsx')
wb = openpyxl.load_workbook('结果表格/所有问题集合.xlsx')
sheet = wb.active
re=0
for the_row in range(1, len(content_list)+1):
    try:
        _ = sheet.cell(row=the_row-re, column = 1, value = content_list[the_row-1])
    except:
        re += 1
        '''
        print(content_list[the_row-1])
        content_list[the_row-1] = content_list[the_row-1].replace('', '')
        content_list[the_row-1] = content_list[the_row-1].replace('', '')
        content_list[the_row-1] = content_list[the_row-1].strip('')
        _ = sheet.cell(row=the_row, column = 1, value = content_list[the_row-1])
        '''

wb.save(filename = '结果表格/所有问题集合.xlsx')
print('已存入所有问题集合.xlsx')
